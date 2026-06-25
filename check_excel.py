"""
Check Excel file for data quality issues
"""

import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('output/Leiloaria_Smart_Grupos_Facebook_COMPLETO.xlsx')
ws = wb['Grupos Facebook']

print('\nSample rows from Excel:')
print('='*120)

# Print header
header = []
for cell in ws[1]:
    header.append(str(cell.value)[:15] if cell.value else "")
print(' | '.join(f'{h:<15}' for h in header))
print('='*120)

# Print first 15 data rows
for row_idx in range(2, min(17, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, 8):
        cell = ws.cell(row_idx, col_idx)
        val = str(cell.value)[:15] if cell.value else ""
        row_data.append(val)
    print(' | '.join(f'{v:<15}' for v in row_data))

print('\n...')
print('='*120)

# Check for issues
print('\n1. CITY NAMES (Truncation Issues):')
print('='*120)
cities_count = defaultdict(int)
cities_list = []

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    city = row[0].value
    if city:
        cities_count[city] += 1
        if city not in cities_list:
            cities_list.append(city)

print(f'Total unique cities: {len(cities_count)}')
print('\nCities found:')
for city in sorted(cities_list):
    print(f'  - {city}')

# Check for wrong population values
print('\n\n2. POPULATION (HABITANTES) ISSUES:')
print('='*120)

population_issues = defaultdict(list)
for row_idx in range(2, min(50, ws.max_row + 1)):
    city = ws.cell(row_idx, 2).value
    habitantes = ws.cell(row_idx, 7).value

    if city and habitantes and habitantes == 0:
        population_issues[city].append(row_idx)

if population_issues:
    print(f'Found {len(population_issues)} cities with 0 inhabitants:')
    for city, rows in sorted(population_issues.items()):
        print(f'  {city}: {len(rows)} entries with 0 habitantes')
else:
    print('No zero habitantes found (good!)')

# Sample of actual data
print('\n\n3. SAMPLE DATA (First 10 entries):')
print('='*120)
for row_idx in range(2, 12):
    busca = ws.cell(row_idx, 1).value
    cidade = ws.cell(row_idx, 2).value
    tipo = ws.cell(row_idx, 3).value
    id_col = ws.cell(row_idx, 4).value
    url = ws.cell(row_idx, 5).value
    nome = ws.cell(row_idx, 6).value
    habitantes = ws.cell(row_idx, 7).value

    print(f'{busca}: {cidade} ({tipo}) - Abitanti: {habitantes}')
    if url:
        print(f'  URL: {str(url)[:60]}...')
    print(f'  Nome: {nome}')
    print()

print('='*120)
