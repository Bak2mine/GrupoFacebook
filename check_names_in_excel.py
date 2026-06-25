"""
Check group names in the generated Excel file
"""

import openpyxl

wb = openpyxl.load_workbook('output/Leiloaria_Smart_Grupos_Facebook_COMPLETO.xlsx')
ws = wb['Grupos Facebook']

print('\nSample of group names from Excel:')
print('='*80)

seen_names = set()
for row_idx in range(2, min(50, ws.max_row + 1)):
    nome = ws.cell(row_idx, 6).value
    if nome and nome not in seen_names:
        seen_names.add(nome)
        print(f'  - {nome}')
        if len(seen_names) >= 15:
            break

print()
print('='*80)
print('Checking for private group entries:')
private_count = 0
for row_idx in range(2, ws.max_row + 1):
    nome = ws.cell(row_idx, 6).value
    if nome == 'Grupo Privado':
        private_count += 1

total_rows = ws.max_row - 1
print(f'Total rows: {total_rows}')
print(f'Rows with "Grupo Privado": {private_count}')
print(f'Rows with real names: {total_rows - private_count}')
print(f'Success rate: {100*(total_rows-private_count)/total_rows:.1f}%')
