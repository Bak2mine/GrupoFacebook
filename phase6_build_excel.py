"""
Phase 6: Generate final Excel report with all group data
Consolidates results from all previous phases
"""

import json
import logging
from pathlib import Path
from typing import List, Dict
from config import (
    DATA_DIR, FINAL_EXCEL, UNIQUE_IDS_JSON, GROUP_NAMES_JSON,
    CITY_POPULATION, BAIRRO_POPULATION, EXCEL_COLUMNS, LOG_LEVEL, LOG_FORMAT
)

# Setup logging
logging.basicConfig(level=getattr(logging, LOG_LEVEL), format=LOG_FORMAT)
logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - install with: pip install openpyxl")


class ExcelReportBuilder:
    """Build final Excel report with all Facebook groups data"""

    def __init__(self):
        self.groups_data = []
        self.search_summary = {}

    def load_data_sources(self):
        """Load all required data files

        Returns:
            Tuple of (groups_with_ids, group_names, searches)
        """
        logger.info("Loading data from all phases...")

        # Load group names
        group_names = {}
        if Path(GROUP_NAMES_JSON).exists():
            with open(GROUP_NAMES_JSON, 'r', encoding='utf-8') as f:
                group_names = json.load(f)
            logger.info(f"Loaded {len(group_names)} group names")
        else:
            logger.warning(f"Group names file not found: {GROUP_NAMES_JSON}")

        # Load searches with group IDs
        searches = []
        searches_file = DATA_DIR / "searches_with_group_ids.json"
        if Path(searches_file).exists():
            with open(searches_file, 'r', encoding='utf-8') as f:
                searches = json.load(f)
            logger.info(f"Loaded {len(searches)} searches")
        else:
            logger.warning(f"Searches file not found: {searches_file}")

        return group_names, searches

    def build_groups_dataset(self, group_names: Dict, searches: List[Dict]) -> List[Dict]:
        """Build dataset for groups sheet

        Args:
            group_names: Mapping of group_id -> name
            searches: List of searches with group IDs

        Returns:
            List of group records
        """
        logger.info("Building groups dataset...")

        groups_data = []

        for search in searches:
            city = search.get('city', 'Unknown')
            state = search.get('state', '')
            search_term = search.get('search_term', '')
            search_type = search.get('type', 'city')
            group_ids = search.get('group_ids', [])
            group_details = search.get('group_details', [])

            # Get population (use bairro if available, otherwise city)
            bairro = search.get('bairro')
            if bairro and bairro in BAIRRO_POPULATION:
                population = BAIRRO_POPULATION[bairro]
            else:
                population = CITY_POPULATION.get(city, 0)

            # Build a map of group_id -> name from group_details
            detail_names = {str(g.get('id')): g.get('name', 'Grupo Privado') for g in group_details}

            for group_id in group_ids:
                group_id_str = str(group_id)

                # Try to use name from group_details first (scraped from search results)
                # Fall back to group_names.json for names from Phase 5 navigation
                # Finally fall back to "Grupo Privado"
                if group_id_str in detail_names:
                    group_name = detail_names[group_id_str]
                else:
                    group_name = group_names.get(group_id_str, 'Grupo Privado')

                # Skip if name is just "Facebook" (private group)
                if group_name.lower() == 'facebook':
                    group_name = 'Grupo Privado'

                # Build Facebook URL
                fb_url = f"https://www.facebook.com/groups/{group_id}"

                group_record = {
                    'Busca': search_term,
                    'Cidade': city,
                    'Tipo': search_type,
                    'ID': str(group_id),
                    'URL': fb_url,
                    'Nome do Grupo': group_name,
                    'Habitantes': population
                }

                groups_data.append(group_record)

        logger.info(f"Built dataset with {len(groups_data)} group records")
        return groups_data

    def build_summary_dataset(self, groups_data: List[Dict]) -> List[Dict]:
        """Build summary dataset aggregated by search

        Args:
            groups_data: List of group records

        Returns:
            List of summary records (one per search)
        """
        logger.info("Building summary dataset...")

        summary_dict = {}

        for group in groups_data:
            key = (group['Busca'], group['Cidade'], group['Tipo'])
            if key not in summary_dict:
                summary_dict[key] = {
                    'Busca': group['Busca'],
                    'Cidade': group['Cidade'],
                    'Tipo': group['Tipo'],
                    'Qtd. Grupos': 0,
                    'Habitantes': group['Habitantes']
                }
            summary_dict[key]['Qtd. Grupos'] += 1

        summary_data = list(summary_dict.values())
        logger.info(f"Built summary with {len(summary_data)} unique searches")
        return summary_data

    def create_excel_workbook(self, groups_data: List[Dict], summary_data: List[Dict]):
        """Create Excel workbook with data

        Args:
            groups_data: List of group records
            summary_data: List of summary records

        Returns:
            Workbook object
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl required for Excel generation")
            return None

        from openpyxl import Workbook

        logger.info("Creating Excel workbook...")

        wb = Workbook()
        ws_groups = wb.active
        ws_groups.title = "Grupos Facebook"

        ws_summary = wb.create_sheet("Resumo por Busca")

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Populate groups sheet
        logger.info("Populating groups sheet...")
        self._populate_sheet(ws_groups, groups_data, "Grupos Facebook",
                            header_font, header_fill, header_alignment, border)

        # Populate summary sheet
        logger.info("Populating summary sheet...")
        self._populate_sheet(ws_summary, summary_data, "Resumo por Busca",
                            header_font, header_fill, header_alignment, border)

        logger.info("Excel workbook created successfully")
        return wb

    def _populate_sheet(self, worksheet, data: List[Dict], sheet_name: str,
                       header_font, header_fill, header_alignment, border):
        """Populate a worksheet with data

        Args:
            worksheet: openpyxl worksheet
            data: List of records
            sheet_name: Name of sheet (for logging)
            header_font: Font for headers
            header_fill: Fill color for headers
            header_alignment: Alignment for headers
            border: Border style
        """
        if not data:
            logger.warning(f"No data for sheet: {sheet_name}")
            return

        # Get columns from first record
        columns = list(data[0].keys())

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Write data rows
        for row_num, record in enumerate(data, 2):
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                value = record.get(column_title, '')

                # Format special columns
                if column_title == 'URL':
                    cell.value = value
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")
                elif column_title == 'Habitantes':
                    cell.value = value
                    cell.number_format = '#,##0'
                elif 'Qtd' in column_title:
                    cell.value = value
                    cell.number_format = '0'
                else:
                    cell.value = value

                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Highlight private groups
                if column_title == 'Nome do Grupo' and value == 'Grupo Privado':
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Adjust column widths
        worksheet.column_dimensions['A'].width = 40  # Busca (wider for full bairro/city names)
        worksheet.column_dimensions['B'].width = 20  # Cidade
        worksheet.column_dimensions['C'].width = 12  # Tipo
        worksheet.column_dimensions['D'].width = 18  # ID
        worksheet.column_dimensions['E'].width = 40  # URL
        worksheet.column_dimensions['F'].width = 30  # Nome do Grupo
        worksheet.column_dimensions['G'].width = 15  # Habitantes

        # Freeze header row
        worksheet.freeze_panes = "A2"

        logger.info(f"Populated {sheet_name} with {len(data)} rows")

    def save_excel(self, workbook, output_file):
        """Save workbook to file

        Args:
            workbook: openpyxl workbook
            output_file: Path to save Excel file
        """
        logger.info(f"Saving Excel file to {output_file}...")

        output_file.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(output_file))

        logger.info(f"Excel file saved: {output_file}")
        logger.info(f"File size: {output_file.stat().st_size / (1024*1024):.2f} MB")

    def run(self):
        """Main pipeline

        Returns:
            Path to generated Excel file
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl library required. Install with: pip install openpyxl")
            return None

        # Load data from all phases
        group_names, searches = self.load_data_sources()

        if not searches:
            logger.error("No search data found. Ensure phases 1-4 are complete.")
            return None

        # Build datasets
        groups_data = self.build_groups_dataset(group_names, searches)
        summary_data = self.build_summary_dataset(groups_data)

        # Create and save Excel
        workbook = self.create_excel_workbook(groups_data, summary_data)
        if workbook:
            self.save_excel(workbook, FINAL_EXCEL)

        # Print statistics
        self.print_statistics(groups_data, summary_data)

        logger.info("Phase 6 complete!")
        return FINAL_EXCEL

    def print_statistics(self, groups_data: List[Dict], summary_data: List[Dict]):
        """Print final statistics

        Args:
            groups_data: List of group records
            summary_data: List of summary records
        """
        logger.info("\n" + "="*70)
        logger.info("FINAL REPORT STATISTICS")
        logger.info("="*70)
        logger.info(f"Total unique groups: {len(groups_data)}")
        logger.info(f"Total searches: {len(summary_data)}")
        logger.info(f"Excel output: {FINAL_EXCEL}")

        # Count private groups
        private_count = sum(1 for g in groups_data if g['Nome do Grupo'] == 'Grupo Privado')
        if private_count > 0:
            logger.info(f"Private groups (not accessible): {private_count}")

        # Sum population
        total_pop = sum(g['Habitantes'] for g in groups_data)
        logger.info(f"Total population covered: {total_pop:,}")


def main():
    """Execute Phase 6"""
    import argparse

    parser = argparse.ArgumentParser(description='Phase 6: Generate Excel report')

    args = parser.parse_args()

    builder = ExcelReportBuilder()
    excel_file = builder.run()

    if excel_file:
        logger.info(f"\nExcel report generated: {excel_file}")


if __name__ == '__main__':
    main()
